"""
Generic XLSX export router.
클라이언트에서 이미 계산한 표 데이터(headers + rows)를 받아
openpyxl로 .xlsx 파일을 만들어 반환한다. (대시보드/달력/상세조회/소재목록 공용)
"""

import io
from urllib.parse import quote

from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

router = APIRouter(prefix="/api/export", tags=["export"])

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class XlsxRequest(BaseModel):
    filename: str = "export.xlsx"
    sheet_name: str = "Sheet1"
    headers: list[str] = []
    rows: list[list] = []


@router.post("/xlsx")
def export_xlsx(body: XlsxRequest) -> StreamingResponse:
    """headers + rows 데이터를 .xlsx 파일로 생성해 다운로드."""
    wb = Workbook()
    ws = wb.active
    # 시트명은 31자 제한, 빈 값 방지
    ws.title = (body.sheet_name or "Sheet1")[:31]

    if body.headers:
        ws.append(body.headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in body.rows:
        ws.append(list(row))

    # 열 너비 자동 조정 (헤더/셀 최대 길이 기준, 한글 가중치 포함)
    for col_idx, _ in enumerate(body.headers, start=1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, values_only=True):
            for v in cell:
                if v is None:
                    continue
                # 한글은 대략 2배 폭으로 계산
                length = sum(2 if ord(ch) > 0x1100 else 1 for ch in str(v))
                max_len = max(max_len, length)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fn = quote(body.filename)
    return StreamingResponse(
        buf,
        media_type=_XLSX_MEDIA,
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{fn}"},
    )
