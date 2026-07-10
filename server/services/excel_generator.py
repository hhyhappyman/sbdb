"""
Excel generator — 공익/재난 방송 월별 송출내역 (xlsx).
공익 시트 + 재난 시트를 하나의 워크북에 저장한다.
"""

from datetime import date as date_cls, time as time_cls

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from config import REPORT_EXCEL_DIR

# ── 공통 스타일 ──────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
_HEADER_FONT = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_thin = Side(style="thin", color="AAAAAA")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _num(v):
    """정수면 int로, 아니면 그대로 (엑셀에 30.0 대신 30 표기)."""
    try:
        f = float(v)
        return int(f) if f.is_integer() else round(f, 1)
    except (TypeError, ValueError):
        return v


def _to_time(time_str: str):
    try:
        h, m, s = (int(x) for x in time_str.split(":"))
        return time_cls(h, m, s)
    except (ValueError, AttributeError):
        return time_str


def _style_header(ws, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _write_common_cells(ws, row_idx: int, r: dict):
    """A:날짜 B:요일 C:방송시간 D:소재명 (공통 앞 4열)."""
    d = ws.cell(row=row_idx, column=1, value=date_cls.fromisoformat(r["date"]))
    d.number_format = "yyyy-mm-dd"
    d.alignment = _CENTER
    ws.cell(row=row_idx, column=2, value=r["weekday"]).alignment = _CENTER
    t = ws.cell(row=row_idx, column=3, value=_to_time(r["time"]))
    t.number_format = "h:mm:ss"
    t.alignment = _CENTER
    ws.cell(row=row_idx, column=4, value=r["name"]).alignment = _LEFT


def _write_gongik_sheet(ws, rows: list[dict]):
    headers = ["날짜", "요일", "방송시간", "공익광고명", "초수", "시급",
               "가중치적용유무", "가중치 적용", "가중치미적용"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, r in enumerate(rows, start=2):
        _write_common_cells(ws, i, r)
        ws.cell(row=i, column=5, value=_num(r["duration"])).alignment = _CENTER
        ws.cell(row=i, column=6, value=r["grade"]).alignment = _CENTER
        ws.cell(row=i, column=7, value="O" if r["weighted"] else None).alignment = _CENTER
        ws.cell(row=i, column=8, value=_num(r["weighted_value"])).alignment = _CENTER
        ws.cell(row=i, column=9, value=_num(r["unweighted_value"])).alignment = _CENTER
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = _BORDER

    widths = {"A": 12, "B": 6, "C": 12, "D": 24, "E": 8, "F": 7, "G": 14, "H": 12, "I": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _write_jaenan_sheet(ws, rows: list[dict]):
    headers = ["날짜", "요일", "방송시간", "공익광고명", "분", "초", "초수(총)",
               "가중치적용유무", "가중치 적용", "가중치미적용"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for i, r in enumerate(rows, start=2):
        _write_common_cells(ws, i, r)
        dur = int(r["duration"] or 0)
        minutes = dur // 60
        secs = dur % 60
        ws.cell(row=i, column=5, value=minutes if minutes > 0 else None).alignment = _CENTER
        ws.cell(row=i, column=6, value=secs if secs > 0 else None).alignment = _CENTER
        ws.cell(row=i, column=7, value=dur).alignment = _CENTER
        ws.cell(row=i, column=8, value="O" if r["weighted"] else None).alignment = _CENTER
        ws.cell(row=i, column=9, value=_num(r["weighted_value"])).alignment = _CENTER
        ws.cell(row=i, column=10, value=_num(r["unweighted_value"])).alignment = _CENTER
        for c in range(1, len(headers) + 1):
            ws.cell(row=i, column=c).border = _BORDER

    widths = {"A": 12, "B": 6, "C": 12, "D": 24, "E": 6, "F": 6, "G": 9,
              "H": 14, "I": 12, "J": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def generate_gongik_jaenan_xlsx(year: int, month: int, data: dict) -> str:
    """
    공익/재난 월별 송출내역 xlsx 생성. 반환: 파일 경로.
    시트: '{month}월(공익)', '{month}월(재난)'.
    """
    wb = Workbook()
    ws_g = wb.active
    ws_g.title = f"{month}월(공익)"
    _write_gongik_sheet(ws_g, data["campaign"])

    ws_j = wb.create_sheet(f"{month}월(재난)")
    _write_jaenan_sheet(ws_j, data["disaster"])

    out_path = str(REPORT_EXCEL_DIR / f"gongik_jaenan_{year}{month:02d}.xlsx")
    wb.save(out_path)
    return out_path
